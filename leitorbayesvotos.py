"""
from openpyxl import load_workbook
caminho = 'C:\Users\Mateus\Documents\IA\housevotes84adpAD.xlsx'
arquivo_excel = load_workbook(caminho)
# Utililizando em índices o nome das células como em um dicionário
c1 = planilha1['C1']
# Imprime o valor da célula C1
print(c1.value)
# Utilizando o método cell
a1 = planilha1.cell(column=1, row=1)
#Imprime o valor da célula a1
print(a1.value) 

max_linha = planilha1.max_row
max_coluna = planilha1.max_column
for i in range(1, max_linha + 1):
    for j in range(1, max_coluna + 1):
        print(planilha1.cell(row=i, column=j).value, end=" - ")


Mateus Felipe Ribeiro - Ra:17243526
"""
# Import stuff we will be needing
import tensorflow as tf
# import tensorflow_probability as tfp
# tfd = tfp.distributions
import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import accuracy_score
from sklearn import datasets, model_selection
import pandas as pd


# Load the dataset
data = pd.read_csv('housevotes84adpAD.csv')
data.replace('N', 0, inplace=True)
data.replace('S', 1, inplace=True)
data.replace('R', 0, inplace=True)
data.replace('D', 1, inplace=True)
data = np.array(data.values.tolist())

# Use only the first two features: sepal length and width
X = data[:, 1:]
Y = data[:, :1].reshape(data.shape[0])

# Randomly shuffle the data and make train and test splits
x_train, x_test, y_train, y_test = model_selection.train_test_split(X, Y, test_size=0.2)

from sklearn.preprocessing import StandardScaler
sc = StandardScaler()
x_train = sc.fit_transform(x_train)
x_test = sc.transform(x_test)

# Training the Naive Bayes model on the Training set
from sklearn.naive_bayes import GaussianNB
classifier = GaussianNB()
classifier.fit(x_train, y_train)

# Predicting the Test set results
y_pred = classifier.predict(x_test)

# Making the Confusion Matrix
from sklearn.metrics import confusion_matrix, accuracy_score
ac = accuracy_score(y_test,y_pred)
cm = confusion_matrix(y_test, y_pred)

print(ac)
print(cm)